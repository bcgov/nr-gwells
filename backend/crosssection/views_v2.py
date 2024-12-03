"""
    Licensed under the Apache License, Version 2.0 (the "License");
    you may not use this file except in compliance with the License.
    You may obtain a copy of the License at

        http://www.apache.org/licenses/LICENSE-2.0

    Unless required by applicable law or agreed to in writing, software
    distributed under the License is distributed on an "AS IS" BASIS,
    WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
    See the License for the specific language governing permissions and
    limitations under the License.
"""

import logging

from django.db import transaction
from django.utils import timezone
from django.http import FileResponse, StreamingHttpResponse
from django.contrib.gis.db.models.functions import Distance
from django.contrib.gis.geos import GEOSException, GEOSGeometry
from django.contrib.gis.gdal import GDALException
from django.db.models.functions import Cast, Lower
from django.db.models import FloatField, Q, Case, When, F, Value, DateField

from rest_framework import status, filters
from rest_framework.exceptions import PermissionDenied, NotFound, ValidationError
from rest_framework.generics import ListAPIView
from rest_framework.response import Response
from gwells.roles import WELLS_VIEWER_ROLE, WELLS_EDIT_ROLE
from gwells.pagination import apiLimitedPagination, APILimitOffsetPagination
from gwells.geojson import GeoJSONIterator

from wells.filters import (
    BoundingBoxFilterBackend,
    WellListFilterBackend,
    WellListOrderingFilter,
    GeometryFilterBackend,
    RadiusFilterBackend,
    WellQaQcFilterBackend
)
from wells.models import Well, WellAttachment, \
  WELL_STATUS_CODE_ALTERATION, WELL_STATUS_CODE_CONSTRUCTION, WELL_STATUS_CODE_DECOMMISSION
from wells.serializers_v2 import (
    WellLocationSerializerV2,
    WellVerticalAquiferExtentSerializerV2,
    WellListSerializerV2,
    WellListAdminSerializerV2,
    WellExportSerializerV2,
    WellExportAdminSerializerV2,
    WellSubsurfaceSerializer,
    WellDetailSerializer,
    MislocatedWellsSerializer,
    CrossReferencingSerializer,
    RecordComplianceSerializer
)
from wells.permissions import WellsEditOrReadOnly, WellsIDIREditOrReadOnly
from wells.renderers import WellListCSVRenderer, WellListExcelRenderer

from aquifers.models import (
    Aquifer,
    VerticalAquiferExtent,
    VerticalAquiferExtentsHistory
)
from aquifers.permissions import HasAquiferEditRole
from wells.views import WellDetail as WellDetailV1
from wells.constants import MAX_EXPORT_COUNT, MAX_LOCATION_COUNT

logger = logging.getLogger(__name__)


class WellLocationListV2APIView(ListAPIView):
    """ Returns well locations for a given search.

        get:
        Returns a list of wells with locations only.
    """
    permission_classes = (WellsEditOrReadOnly,)
    model = Well
    pagination_class = apiLimitedPagination(MAX_LOCATION_COUNT)

    # Allow searching on name fields, names of related companies, etc.
    filter_backends = (WellListFilterBackend, BoundingBoxFilterBackend,
                       filters.SearchFilter, WellListOrderingFilter, GeometryFilterBackend)
    ordering = ('well_tag_number',)

    search_fields = ('well_tag_number', 'identification_plate_number',
                     'street_address', 'city', 'owner_full_name', 'ems')

    TOO_MANY_ERROR_MESSAGE = "Too many wells to display on map. Please zoom in or change your search criteria."

    def get_serializer_class(self):
        return WellLocationSerializerV2

    def get_queryset(self):
        """ Excludes Unpublished wells for users without edit permissions """
        if self.request.user.groups.filter(name=WELLS_EDIT_ROLE).exists():
            qs = Well.objects.all()
        else:
            qs = Well.objects.all().exclude(well_publication_status='Unpublished')

        # check to see if we should filter wells by which ones intersect an aquifer
        intersects_aquifer_id = self.request.query_params.get('intersects_aquifer_id', None)
        if intersects_aquifer_id:
            aquifer = Aquifer.objects.filter(aquifer_id=int(intersects_aquifer_id)).first()

            if not aquifer:
                raise NotFound(f'Unknown aquifer {intersects_aquifer_id}')

            if not aquifer.geom:
                # if the aquifer has no/null geometry, it might be an aquifer
                # that the business area has created but has not delineated an area
                # for (for example, the special "holding" aquifer 1143).
                qs = qs.none()
                
            else:
                # Find wells that intersect this simplified aquifer polygon (excluding wells
                # with null geom)
                qs = qs.exclude(geom=None)
                qs = qs.filter(geom__intersects=aquifer.geom)

        well_tag_numbers = self.request.query_params.get('well_tag_numbers', '')
        if well_tag_numbers:
            well_tag_numbers = well_tag_numbers.split(',')
            qs = qs.filter(well_tag_number__in=well_tag_numbers)

        return qs

    def get(self, request, *args, **kwargs):
        """
        Returns geojson if requested, otherwise handles request as normal.
        """

        geojson_requested = self.request.query_params.get('geojson') == 'true'

        # if geojson requested, create a query that returns each well's geometry as GeoJSON
        # so that we can easily create a FeatureCollection.
        # This might be more performant in the database using json_agg and ST_AsGeoJSON
        # vs creating geojson Features here in Python.
        if geojson_requested:
            return self.geoJSONResponse()

        return super().get(request)

    def geoJSONResponse(self):
        """
        Returns a streaming GeoJSON HTTP response of the searched wells
        """
        qs = self.get_queryset()
        qs = qs.exclude(geom=None)

        fields = [
            "geom",
            "well_tag_number",
            "identification_plate_number",
            "street_address",
            "city",
            "well_status",
            "artesian_conditions",
            "storativity",
            "transmissivity",
            "hydraulic_conductivity"
        ]

        locations = self.filter_queryset(qs)

        # If the user can edit wells then we can add the `is_published` property to the response
        if self.request.user.groups.filter(name=WELLS_EDIT_ROLE).exists():
            locations = locations.extra(select={'is_published': "well_publication_status_code = 'Published'"})
            fields.append("is_published")

        locations = locations.values(*fields)
        locations = list(locations[:MAX_LOCATION_COUNT + 1])

        # return a 403 response if there are too many wells to display
        if len(locations) > MAX_LOCATION_COUNT:
            raise PermissionDenied(self.TOO_MANY_ERROR_MESSAGE)

        # turn the list of locations into a generator so the GeoJSONIterator can use it
        locations_iter = (location for location in locations)
        iterator = GeoJSONIterator(locations_iter)

        return StreamingHttpResponse(iterator, content_type="application/json")


class WellAquiferListV2APIView(ListAPIView):
    """
    Returns a list of aquifers with depth information for a well.
    """
    permission_classes = (HasAquiferEditRole,)
    ordering = ('start',)
    serializer_class = WellVerticalAquiferExtentSerializerV2
    pagination_class = None

    def get_queryset(self):
        """
        Excludes Aquifer 3D points that relate to unpublished wells for users without edit permissions
        """
        well = self.get_well()

        qs = VerticalAquiferExtent.objects.filter(well=well).select_related('aquifer')

        if not self.request.user.groups.filter(name=WELLS_EDIT_ROLE).exists():
            qs = qs.exclude(well__well_publication_status='Unpublished')

        return qs

    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        username = self.request.user.profile.username
        timestamp = timezone.now()

        # we expect a list
        if not isinstance(request.data, list):
            raise NotFound()

        # get the well and 404 if it doesn't exist
        well = self.get_well()
        max_depth = float('-inf')
        ids = []
        items = []
        errors = []
        has_errors = False
        for item in request.data:  # go through each vertical aquifer extent
            item['well_tag_number'] = well.well_tag_number

            vertical_aquifer_extent = None
            vae_id = item.get('id', None)
            if vae_id:  # has an id - then it must be an existing one
                vertical_aquifer_extent = VerticalAquiferExtent.objects.get(pk=vae_id)

            serializer = WellVerticalAquiferExtentSerializerV2(instance=vertical_aquifer_extent,
                                                               data=item)
            serializer_errors = {}
            if serializer.is_valid():
                # add user audit information
                serializer.validated_data['update_user'] = username
                serializer.validated_data['update_date'] = timestamp
                if not vertical_aquifer_extent:
                    serializer.validated_data['create_user'] = username
                    serializer.validated_data['create_date'] = timestamp

                if self.has_changed(vertical_aquifer_extent, serializer.validated_data):
                    vertical_aquifer_extent = serializer.save()

                # keep track existing ids and any newly added IDs
                ids.append(vertical_aquifer_extent.id)
                items.append(serializer.data)
            else:
                serializer_errors = serializer.errors
                has_errors = True

            if vertical_aquifer_extent is not None:
                self.log_history(vertical_aquifer_extent, username, timestamp)

                if vertical_aquifer_extent.start < max_depth:
                    has_errors = True
                    serializer_errors.setdefault('start', []) \
                        .append('Start depth overlaps with another')

                max_depth = vertical_aquifer_extent.end

            errors.append(serializer_errors)  # always add to keep the index correct for web app

        # roll back on errors and undo any changes
        if has_errors:
            transaction.set_rollback(True)
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        # delete any ids not in the POST-ed list
        self.get_queryset().exclude(id__in=ids).delete()

        return Response(items, status=status.HTTP_201_CREATED)

    def get_well(self):
        well_tag_number = int(self.kwargs['well_tag_number'])
        try:
            return Well.objects.get(pk=well_tag_number)
        except Exception:
            raise NotFound(f'Well {well_tag_number} could not be found')

    def has_changed(self, existing_vertical_aquifer_extent, new_data):
        if existing_vertical_aquifer_extent is None:
            return True

        if existing_vertical_aquifer_extent.start != new_data['start']:
            return True

        if existing_vertical_aquifer_extent.end != new_data['end']:
            return True

        if existing_vertical_aquifer_extent.aquifer_id != new_data['aquifer_id']:
            return True

        if existing_vertical_aquifer_extent.geom and new_data['geom']:
            if existing_vertical_aquifer_extent.geom.x != new_data['geom'].x:
                return True

            if existing_vertical_aquifer_extent.geom.y != new_data['geom'].y:
                return True
        else:
            return True

        return False

    def log_history(self, vertical_aquifer_extent, username, timestamp):
        # Whenever a VerticalAquiferExtent is saved - insert a copy of the data into the
        # vertical_aquifer_extents_history table
        VerticalAquiferExtentsHistory.objects.create(
            create_user=username,
            create_date=timestamp,
            update_user=username,
            update_date=timestamp,
            well_tag_number=vertical_aquifer_extent.well_id,
            aquifer_id=vertical_aquifer_extent.aquifer_id,
            geom=vertical_aquifer_extent.geom,
            start=vertical_aquifer_extent.start,
            end=vertical_aquifer_extent.end
        )


class WellListAPIViewV2(ListAPIView):
    """List and create wells

    get:
    Returns a list of wells.
    """

    permission_classes = (WellsEditOrReadOnly,)
    model = Well
    pagination_class = APILimitOffsetPagination

    filter_backends = (WellListFilterBackend, BoundingBoxFilterBackend,
                       filters.SearchFilter, WellListOrderingFilter, GeometryFilterBackend)
    ordering = ('well_tag_number',)
    search_fields = ('well_tag_number', 'identification_plate_number',
                     'street_address', 'city', 'owner_full_name')
    default_limit = 10

    def get_serializer_class(self):
        """Returns a different serializer class for admin users."""
        serializer_class = WellListSerializerV2
        if (self.request.user and self.request.user.is_authenticated and
                self.request.user.groups.filter(name=WELLS_VIEWER_ROLE).exists()):
            serializer_class = WellListAdminSerializerV2

        return serializer_class

    def get_queryset(self):
        """ Excludes Unpublished wells for users without edit permissions """
        if self.request.user.groups.filter(name=WELLS_EDIT_ROLE).exists():
            qs = Well.objects.all()
        else:
            qs = Well.objects.all().exclude(well_publication_status='Unpublished')

        qs = qs \
            .select_related(
                "bcgs_id",
            ).prefetch_related(
                "water_quality_characteristics",
                "drilling_methods",
                "development_methods"
            )

        return qs


class WellExportListAPIViewV2(ListAPIView):
    """Returns CSV or Excel data for wells.
    """
    permission_classes = (WellsEditOrReadOnly,)
    model = Well

    # Allow searching on name fields, names of related companies, etc.
    filter_backends = (WellListFilterBackend, BoundingBoxFilterBackend,
                       filters.SearchFilter, filters.OrderingFilter)
    ordering = ('well_tag_number',)
    pagination_class = None

    search_fields = ('well_tag_number', 'identification_plate_number',
                     'street_address', 'city', 'owner_full_name')
    renderer_classes = (WellListCSVRenderer, WellListExcelRenderer)

    SELECT_RELATED_OPTIONS = [
        'well_class',
        'well_subclass',
        'well_status',
        'land_district',
        'company_of_person_responsible',
        'ground_elevation_method',
        'surface_seal_material',
        'surface_seal_method',
        'liner_material',
        'screen_intake_method',
        'screen_type',
        'screen_material',
        'screen_opening',
        'screen_bottom',
        'well_yield_unit',
        'observation_well_status',
        'coordinate_acquisition_code',
        'bcgs_id',
        'decommission_method',
        'aquifer',
        'aquifer_lithology',
        'yield_estimation_method',
        'well_disinfected_status',
    ]
    PREFETCH_RELATED_OPTIONS = [
        'development_methods',
        'drilling_methods',
        'water_quality_characteristics',
    ]

    def get_fields(self):
        raw_fields = self.request.query_params.get('fields')
        return raw_fields.split(',') if raw_fields else None

    def get_queryset(self):
        """Excludes unpublished wells for users without edit permissions.
        """
        if self.request.user.groups.filter(name=WELLS_EDIT_ROLE).exists():
            qs = Well.objects.all()
        else:
            qs = Well.objects.all().exclude(well_publication_status='Unpublished')

        included_fields = self.get_fields()

        if included_fields:
            select_relateds = [
                relation for relation in self.SELECT_RELATED_OPTIONS
                if relation in included_fields
            ]
            prefetches = [
                relation for relation in self.PREFETCH_RELATED_OPTIONS
                if relation in included_fields
            ]

            if select_relateds:
                qs = qs.select_related(*select_relateds)
            if prefetches:
                qs = qs.prefetch_related(*prefetches)
        elif included_fields is None:
            # If no fields are passed, then include everything
            qs = qs.select_related(*self.SELECT_RELATED_OPTIONS)
            qs = qs.prefetch_related(*self.PREFETCH_RELATED_OPTIONS)

        return qs

    def get_serializer_class(self):
        """Returns a different serializer class for admin users."""
        serializer_class = WellExportSerializerV2
        if (self.request.user and self.request.user.is_authenticated and
                self.request.user.groups.filter(name=WELLS_VIEWER_ROLE).exists()):
            serializer_class = WellExportAdminSerializerV2

        return serializer_class

    def get_serializer_context(self):
        context = super().get_serializer_context()

        fields = self.get_fields()
        if fields:
            context['fields'] = fields

        return context

    def get_renderer_context(self):
        context = super().get_renderer_context()

        fields = self.get_fields()
        if fields:
            context['header'] = fields

        return context

    def batch_iterator(self, queryset, count, batch_size=200):
        """Batch a queryset into chunks of batch_size, and serialize the results

        Allows iterative processing while taking advantage of prefetching many
        to many relations.
        """
        for offset in range(0, count, batch_size):
            end = min(offset + batch_size, count)
            batch = queryset[offset:end]

            serializer = self.get_serializer(batch, many=True)
            for item in serializer.data:
                yield item

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        count = queryset.count()
        # return an empty response if there are too many wells to display
        if count > MAX_EXPORT_COUNT:
            raise PermissionDenied(
                'Too many wells to export. Please change your search criteria.'
            )
        elif count == 0:
            raise NotFound('No well records could be found.')

        renderer = request.accepted_renderer
        if renderer.format == 'xlsx':
            response_class = FileResponse
        else:
            response_class = StreamingHttpResponse

        context = self.get_renderer_context()
        data_iterator = self.batch_iterator(queryset, count)
        render_result = renderer.render(data_iterator, renderer_context=context)

        response = response_class(render_result, content_type=renderer.media_type)
        response['Content-Disposition'] = 'attachment; filename="search-results.{ext}"'.format(ext=renderer.format)

        return response


class WellSubsurface(ListAPIView):
    """
    This replaces WellScreen with the additional aquifer and lithology info
    
    get:
    Returns well subsurface info within a geometry or a list of wells.
    """

    model = Well
    serializer_class = WellSubsurfaceSerializer
    filter_backends = (GeometryFilterBackend, RadiusFilterBackend)

    def get_queryset(self):
        qs = Well.objects.all() \
            .select_related('intended_water_use', 'aquifer', 'aquifer__material',
                            'aquifer__subtype') \
            .prefetch_related('screen_set')

        if not self.request.user.groups.filter(name=WELLS_EDIT_ROLE).exists():
            qs = qs.exclude(well_publication_status='Unpublished')

        # check if a point was supplied (note: actual filtering will be by
        # the filter_backends classes).  If so, add distances from the point.
        point = self.request.query_params.get('point', None)
        srid = self.request.query_params.get('srid', 4326)
        radius = self.request.query_params.get('radius', None)
        if point and radius:
            try:
                shape = GEOSGeometry(point, srid=int(srid))
                radius = float(radius)
                assert shape.geom_type == 'Point'
            except (ValueError, AssertionError, GDALException, GEOSException):
                raise ValidationError({
                    'point': 'Invalid point geometry. Use geojson geometry or WKT. Example: {"type": "Point", "coordinates": [-123,49]}'
                })
            else:
                qs = qs.annotate(
                    distance=Cast(Distance('geom', shape), output_field=FloatField())
                ).order_by('distance')

        # can also supply a comma separated list of wells
        wells = self.request.query_params.get('wells', None)

        if wells:
            wells = wells.split(',')

            for w in wells:
                if not w.isnumeric():
                    raise ValidationError(detail='Invalid well')

            wells = map(int, wells)
            qs = qs.filter(well_tag_number__in=wells)

        return qs


class WellDetail(WellDetailV1):
    """
    Return well detail.
    This view is open to all, and has no permissions.

    get:
    Returns details for a given well matching the well_tag_number.
    """
    serializer_class = WellDetailSerializer

# QaQc Views

class MislocatedWellsListView(ListAPIView):
    """
    API view to retrieve mislocated wells.
    """
    serializer_class = MislocatedWellsSerializer

    swagger_schema = None
    permission_classes = (WellsIDIREditOrReadOnly,)
    model = Well
    pagination_class = APILimitOffsetPagination

    # Allow searching on name fields, names of related companies, etc.
    filter_backends = (WellListOrderingFilter, WellQaQcFilterBackend,
                       filters.SearchFilter)

    ordering = ('well_tag_number',)

    def get_queryset(self):
        """
        This view should return a list of all mislocated wells
        for the currently authenticated user.
        """
        queryset = Well.objects.all()

        return queryset


class RecordComplianceListView(ListAPIView):
    serializer_class = RecordComplianceSerializer

    swagger_schema = None
    permission_classes = (WellsIDIREditOrReadOnly,)
    model = Well
    pagination_class = APILimitOffsetPagination

    # Allow searching on name fields, names of related companies, etc.
    filter_backends = (WellListOrderingFilter, WellQaQcFilterBackend, 
                       filters.SearchFilter)
    ordering = ('well_tag_number',)

    def get_queryset(self):
        queryset = Well.objects.all()

        return queryset
    

class CrossReferencingListView(ListAPIView):
    serializer_class = CrossReferencingSerializer

    swagger_schema = None
    permission_classes = (WellsIDIREditOrReadOnly,)
    model = Well
    pagination_class = APILimitOffsetPagination

    # Allow searching on name fields, names of related companies, etc.
    filter_backends = (WellListOrderingFilter, WellQaQcFilterBackend,
                       filters.SearchFilter)
    ordering = ('well_tag_number',)

    def get_queryset(self):
        """
        Optionally restricts the returned wells to those that have certain keywords like 'x-ref'd' or 'cross-ref'
        in their internal_comments.
        """
        queryset = Well.objects.filter(cross_referenced=True)

        return queryset

# Download Views for QaQc

class MislocatedWellsDownloadView(WellExportListAPIViewV2):
    permission_classes = (WellsIDIREditOrReadOnly,)
    filter_backends = (WellListOrderingFilter, WellQaQcFilterBackend, filters.SearchFilter)

    def get_queryset(self):
        return Well.objects.all()

    def get_serializer_class(self):
        return MislocatedWellsSerializer
    

class RecordComplianceDownloadView(WellExportListAPIViewV2):
    permission_classes = (WellsIDIREditOrReadOnly,)
    filter_backends = (WellListOrderingFilter, WellQaQcFilterBackend, filters.SearchFilter)

    def get_queryset(self):
        return Well.objects.all()

    def get_serializer_class(self):
        return RecordComplianceSerializer
    

class CrossReferencingDownloadView(WellExportListAPIViewV2):
    permission_classes = (WellsIDIREditOrReadOnly,)
    filter_backends = (WellListOrderingFilter, WellQaQcFilterBackend, filters.SearchFilter)

    def get_queryset(self):
        # Return wells that have been cross-referenced
        return Well.objects.filter(cross_referenced=True)

    def get_serializer_class(self):
        return CrossReferencingSerializer

async def parse_result(res: ClientResponse) -> LineString:
    body = await res.read()
    line = {}

    try:
        line = json.loads(body)
    except TypeError as e:
        logger.error(e)
    except json.JSONDecodeError as e:
        logger.error(e)

    # check if fc looks like a geojson FeatureCollection, and if so,
    # make proper Features out of all the objects
    # if res.status == 200 and line.get("altitude") != None

    return LineString(
        [
            (
                point.get("geometry").get("coordinates")[0],
                point.get("geometry").get("coordinates")[1],
                point.get("altitude")
            ) for point in line
        ]
    )


async def fetch(line: str, session: ClientSession) -> asyncio.Future:
    # asynchronously fetch one URL, expecting a geojson response
    steps = 20
    if not line:
        return []
    url = f"http://geogratis.gc.ca/services/elevation/cdem/profile.json?path={line}&steps={steps}"
    logger.info("external request: %s", url)
    async with session.get(url, raise_for_status=True) as response:
        return await asyncio.ensure_future(parse_result(response))


async def batch_fetch(
        semaphore: asyncio.Semaphore,
        req: str,
        session: ClientSession) -> asyncio.Future:
    # batch_fetch uses a semaphore to make batched requests in parallel
    # (up a limit equal to the size of the semaphore).

    async with semaphore:
        return await fetch(req, session)


async def fetch_all(requests: List[str]) -> asyncio.Future:
    tasks = []
    semaphore = asyncio.Semaphore(10)
    headers = {'accept': 'application/json'}

    async with ClientSession(headers=headers) as session:
        for req in requests:
            task = asyncio.ensure_future(
                batch_fetch(semaphore, req, session))
            tasks.append(task)

        # return the gathered tasks,
        # which will be a list of JSON responses when all requests return.
        return await asyncio.gather(*tasks)


@retry(wait=wait_fixed(2), stop=stop_after_attempt(2))
def fetch_surface_lines(requests: List[str]) -> List[Feature]:
    return asyncio.run(fetch_all(requests))

class ElevationView(View):
    def get(self, request):
        # Fetch request data (e.g., a list of lines from request params)
        requests = request.GET.getlist("lines")  # Assume lines come as a query parameter

        try:
            # Call the async function that fetches elevation data
            result = fetch_surface_lines(requests)
            return JsonResponse({"status": "success", "data": result})

        except Exception as e:
            logger.error(f"Error fetching surface lines: {str(e)}")
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

def get_profile_geojson(line: LineString) -> list:
    # get geojson elevations along a line (GeoGratis - Government of Canada)
    steps = 10
    line = line.wkt
    if not line:
        return []

    resp = requests.get(
        f"http://geogratis.gc.ca/services/elevation/cdem/profile.json?path={line}&steps={steps}"
    )

    return resp.json()


def geojson_to_profile_line(elevations: list) -> LineString:
    # uses GeoGratis (Government of Canada) API to retrieve elevation along a profile
    # line (as a LineString shape object

    profile_line = LineString(
        [
            (
                f.get("geometry").get("coordinates")[0],
                f.get("geometry").get("coordinates")[1],
                f.get("altitude")
            ) for f in elevations
        ]
    )

    return profile_line


def get_profile_line_by_length(db: Session, line: LineString):
    # convert a LineStringZ (3d line) to elevations along the length of the line

    q = """
    SELECT
            ST_Distance(ST_Force2D(geom),
                ST_StartPoint(
                    ST_Transform(
                        ST_SetSRID(
                            ST_GeomFromText(:line),
                            4326
                        ),
                        3005
                    )
                )
            ) as distance_from_origin,
            ST_Z(geom) as elevation
    FROM
        (select
            (
                ST_DumpPoints(
                    ST_Transform(
                        ST_SetSRID(
                            ST_GeomFromText(:line),
                            4326
                        ),
                        3005
                    )
                )
            ).geom
        ) as pts;
    """

    elevation_profile = []

    rows = db.execute(q, {"line": line.wkt})
    for row in rows:
        elevation_profile.append(Elevation(distance_from_origin=row[0], elevation=row[1]))

    return elevation_profile

def get_wells_by_distance(db, search_point, radius):
  # List wells by distance from a point.
    if radius > 10000:
        radius = 10000
    
    # search within a given radius, adding a distance column denoting
    # distance from the centre point in metres
    # geometry columns are cast to geography to use metres as the base unit.
    wells = GroundWaterWells.objects.annotate(
        distance=Distance('geometry', search_point)
    ).filter(
        distance__lte=radius
    ).order_by('distance')
    
    return wells

class WellDrawdownAPIView(APIView):
  # takes a list of WellDrawdown objects and fills in drawdown calculations
    def get(self, request):
        point = Point(float(request.query_params['longitude']), float(request.query_params['latitude']))
        radius = float(request.query_params['radius'])
        
        wells = get_wells_by_distance(request, point, radius)
        return Response(wells, status=status.HTTP_200_OK)




def calculate_available_drawdown(wells: List[WellDrawdown]) -> List[WellDrawdown]:
    """ takes a list of WellDrawdown objects and fills in drawdown calculations """

    for well in wells:
        if well.screen_set:
            # well has a screen set: calculate the top of screen using
            # the screen set supplied by GWELLS.
            well.top_of_screen = calculate_top_of_screen(well.screen_set)

        if well.top_of_screen and well.static_water_level:
            # calculate the difference between the static water level
            # and the top of the screen.  This value indicates the
            # available drawdown. This calculation depends on the reported
            # values available at the time that the well report was filed.
            well.swl_to_screen = well.top_of_screen - well.static_water_level

        if well.finished_well_depth and well.static_water_level:
            # calculate difference between static water level and
            # the finished well depth.  The finished well depth is available
            # on more wells than screen depths are.
            well.swl_to_bottom_of_well = well.finished_well_depth - well.static_water_level

    return wells


def calculate_top_of_screen(screen_set: List[Screen]) -> Optional[float]:
    """ calculates the top of screen from a given screen set
    screen sets come from GWELLS and have a start depth and end depth."""

    top_of_screen = None

    if not screen_set or None in map(lambda x: x.start, screen_set):
        return None

    try:
        top_of_screen = min([x.start for x in screen_set if x.start])
    except ValueError:
        # we expect occasional ValueErrors due to inconsistent screen data.
        # some screens are present in the dataset but do not have start/end values.
        return None
    return top_of_screen


def get_wells_by_aquifer(point, radius, well_tag_numbers=None) -> Dict[Union[int, str], List[WellDrawdown]]:
    """Get wells, grouped by aquifer number"""
    wells = get_wells_with_drawdown(point, radius, well_tag_numbers)

    aquifers = set()

    # Get aquifers
    for well in wells:

        if well.aquifer:
            aquifers.add(well.aquifer.aquifer_id)
        else:
            aquifers.add(None)

    wells_by_aquifer = {}

    for a in aquifers:
        wells_by_aquifer[a if a else ''] = [w for w in wells if
                                            (w.aquifer and w.aquifer.aquifer_id == a) or (
                                                    a is None and not w.aquifer)]
    return wells_by_aquifer


def get_wells_with_drawdown(point, radius, well_tag_numbers=None) -> List[WellDrawdown]:
    """ Find wells near a given point, with a buffer radius,
        or a list of wells (comma separated well tag numbers)
        This function gets wells and their corresponding subsurface data using the GWELLS API
        and then computes the distance of the point to the wells
    """

    if well_tag_numbers is None:
        well_tag_numbers = ''

    wells_results = []

    done = False

    if well_tag_numbers:
        url = f"{GWELLS_API_URL}/api/v2/wells/subsurface?wells={well_tag_numbers}"
    else:
        buffer = create_circle_polygon(point, radius)
        url = f"{GWELLS_API_URL}/api/v2/wells/subsurface?within={buffer.wkt}&limit=100"

    # helpers to prevent unbounded requests
    limit_requests = 100
    i = 0  # this i is for recording extra requests within each chunk, if necessary

    while not done and i < limit_requests:
        logger.info('external request: %s', url)
        resp = requests.get(url)

        i += 1
        # break now if we didn't receive any results.
        results = resp.json().get('results', None)
        if not results:
            done = True
            break

        for well in results:
            # calculate distance from well to click point
            center_point = transform(transform_4326_3005, point)
            well_point = transform(transform_4326_3005, Point(well["longitude"], well["latitude"]))
            distance = center_point.distance(well_point)
            well["distance"] = distance

        # add results to a list.
        wells_results += [WellDrawdown(**well) for well in results]

        # check for a "next" attribute, indicating the next limit/offset combo.
        # when it is null, the pagination is done.
        next_url = resp.json().get('next', None)
        if not next_url:
            done = True
        url = next_url

    # return zero results if an error occurred or we did not successfully get all the results.
    # (avoid returning incomplete data)
    if not done:
        return []

    wells = calculate_available_drawdown(wells_results)

    return wells


def merge_wells_datasources(wells: list, wells_with_distances: object) -> List[WellDrawdown]:
    """
    Merges a list of well details (from GWELLS), with a key/value dict of wells: distance (m)
    to create a list of WellDrawdown data.
    e.g. combines:
        {
            123: 50,
            124: 55
        }
    with:
        [
            {
                well_tag_number: 123,
                static_water_level: 12
            },
            {
                well_tag_number: 124,
                static_water_level: 12
            }
        ]
    to create:
        [
            {
                well_tag_number: 123,
                static_water_level: 12,
                distance: 50
            },
            {
                well_tag_number: 124,
                static_water_level: 12,
                distance: 55
            }
        ]
    """

    well_map = {}

    # make a dict with keys being the well tag numbers
    for well in wells:
        well_map[str(well.pop('well_tag_number'))] = well

    # create WellDrawdown data objects for every well we found nearby.
    # The last argument to WellDrawdown() is
    # the supplemental data that comes from GWELLS for each well.
    return calculate_available_drawdown([
        WellDrawdown(
            well_tag_number=well[0],
            distance=well[1],
            **well_map.get(str(well[0]).lstrip('0'), {})
        )
        for well in wells_with_distances])


def create_circle_polygon(point: Point, radius: float):
    point = transform(transform_4326_3005, point)
    circle = point.buffer(radius)
    return transform(transform_3005_4326, circle)


def create_line_buffer(line: LineString, radius: float):
    line = transform(transform_4326_3005, line)
    buf = line.buffer(radius, cap_style=CAP_STYLE.flat,
                      join_style=JOIN_STYLE.round)
    return transform(transform_3005_4326, buf)


def get_line_buffer_polygon(line: LineString, radius: float):
    """ returns a buffer area around a LineString. """
    return func.ST_Transform(func.ST_Buffer(
        func.St_Transform(
            func.ST_GeomFromText(line.wkt, 4326),
            3005
        ),
        radius,
        'endcap=flat join=round'
    ), 4326)


def get_parallel_line_offset(db: Session, line: LineString, radius: float):
    """ returns a parallel line perpendicular to a LineString. """
    return db.query(func.ST_AsText(func.ST_Transform(func.ST_OffsetCurve(
        func.St_Transform(
            func.ST_GeomFromText(line.wkt, 4326),
            3005
        ),
        radius
    ), 4326))).first()


def distance_along_line(line: LineString, point: Point, srid=4326):
    """
    calculates the distance that `point` is along `line`. Note that
    this is the distance along the line, not from the beginning of the line
    to the point.
    """

    if srid == 4326:
        # transform to BC Albers, which has a base unit of metres
        point = transform(transform_4326_3005, point)
        line = transform(transform_4326_3005, line)

    elif srid != 3005:
        raise ValueError("SRID must be either 4326 or 3005")

    # note.  shapely's geom.distance calculates distance on a 2d plane
    c = point.distance(line.interpolate(0))
    b = point.distance(line)
    return math.sqrt(abs(c ** 2 - b ** 2))


def elevation_along_line(profile, distance):
    """ returns the elevation at `distance` metres along LineString Z `profile` """
    profile = transform(transform_4326_3005, profile)
    return profile.interpolate(distance).z


def get_wells_along_line(db: Session, profile: LineString, radius: float):
    """ returns wells along a given line, including wells that are within a buffer
        determined by `radius` (m).
        `radius` creates a buffer area next to the line that does not include any area
        behind or beyond the start/end of the drawn line. The wells are ordered
        by the distance from the origin (i.e. the beginning of the line, measured
        along the axis).
    """
    buf = create_line_buffer(profile, radius)

    req = ExternalAPIRequest(
        url=f"{GWELLS_API_URL}/api/v2/wells/subsurface",
        q={
            "within": buf.wkt,
            "limit": 100
        },
        layer="gwells"
    )
    feature_collection = fetch_geojson_features([req])[0].geojson

    wells_results = []

    for well in feature_collection.features:
        line = LineString([coords[:2] for coords in list(profile.coords)])
        point = Point(shape(well.geometry))

        shortest_line = distance_from_line(line, point)
        distance = distance_along_line(line, point)
        compass_direction = compass_direction_point_to_line(line, point)

        # Separate the well aquifer info from the feature info
        well_aquifer = well.properties.pop('aquifer', None)

        # Add (flattened) aquifer into feature info
        well.properties['aquifer'] = well_aquifer.get('aquifer_id') if well_aquifer else None

        well.properties['distance_from_line'] = shortest_line
        well.properties['compass_direction'] = compass_direction

        # Remove lithologydescription_set from well properties as it's not formatted properly
        well.properties.pop('lithologydescription_set')

        # load screen data from the geojson response
        screenset = well.properties.get('screen_set', '')
        screenset = json.loads(screenset)

        well_data = {
            "well_tag_number": well.properties['well_tag_number'],
            "finished_well_depth": float(well.properties['finished_well_depth']) * 0.3048
            if well.properties['finished_well_depth'] else None,
            "water_depth": float(well.properties['static_water_level']) * 0.3048 if well.properties[
                'static_water_level'] else None,
            "distance_from_origin": distance,
            "ground_elevation_from_dem": elevation_along_line(profile, distance),
            "aquifer": well_aquifer,
            "aquifer_lithology": well.properties['aquifer_lithology'],
            "feature": well,
            "screen_set": screenset
        }

        wells_results.append(well_data)

    return wells_results


def get_waterbodies_along_line(section_line: LineString, profile: LineString):
    """ retrieves streams that cross the cross section profile line """

    line_3005 = transform(transform_4326_3005, section_line)

    streams_layer = "WHSE_BASEMAPPING.FWA_STREAM_NETWORKS_SP"
    lakes_layer = "WHSE_BASEMAPPING.FWA_LAKES_POLY"

    cql_filter = f"""INTERSECTS(GEOMETRY, {line_3005.wkt})"""

    intersecting_lakes = databc_feature_search(
        lakes_layer, cql_filter=cql_filter)
    intersecting_streams = databc_feature_search(
        streams_layer, cql_filter=cql_filter)

    stream_features = []
    lake_features = []

    # create a MultiPolygon of all the lake geometries.
    # this will be used to check if a stream intersection falls inside a lake
    # (lake names will supersede stream names inside lakes)
    lake_polygons = []
    for lake in intersecting_lakes.features:
        geom = shape(lake.geometry)
        if isinstance(geom, MultiPolygon):
            lake_polygons = lake_polygons + [poly for poly in geom]
        elif isinstance(geom, Polygon):
            lake_polygons.append(geom)

    lakes_multipoly_shape = MultiPolygon(lake_polygons)

    # convert each intersecting stream into a Point or MultiPoint using .intersection().
    # check each point of intersection to make sure it doesn't lie on a lake (stream lines in
    # the Freshwater Atlas extend through lakes, but when we are over a lake, we want the lake
    # name not the stream name).
    # the elevation for points comes from the Freshwater Atlas,
    # so it's possible it could be slightly off
    # the CDEM value from the Canada GeoGratis DEM API.
    for stream in intersecting_streams.features:
        intersecting_points = line_3005.intersection(shape(stream.geometry))

        # the intersection may either be a MultiPoint (which is iterable),
        # or a single Point instance (not iterable). If not iterable, convert
        # to a list of 1 Point.
        if isinstance(intersecting_points, Point):
            intersecting_points = [intersecting_points]

        for point in intersecting_points:
            if point.intersects(lakes_multipoly_shape):
                # skip so that we can defer to the lake's name
                continue

            distance = distance_along_line(
                LineString([coords[:2] for coords in list(line_3005.coords)]),
                point,
                srid=3005
            )
            stream_data = {
                "name": stream.properties['GNIS_NAME'] or "Unnamed Stream",
                "distance": distance,
                "elevation": point.z,
                "geometry": transform(transform_3005_4326, point)
            }
            stream_features.append(stream_data)

    # for lakes, use a representative point (using the centroid).
    # Lakes don't come with an elevation, so the elevation uses the profile
    # retrieved from the GeoGratis CDEM API.
    for lake in intersecting_lakes.features:
        intersecting_points = line_3005.intersection(shape(lake.geometry))

        if isinstance(intersecting_points, LineString):
            intersecting_points = [intersecting_points]

        for line in intersecting_points:
            point = line.centroid
            distance = distance_along_line(
                LineString([coords[:2] for coords in list(line_3005.coords)]),
                point,
                srid=3005
            )
            lake_data = {
                "name": lake.properties['GNIS_NAME_1'] or f"Unnamed Lake",
                "distance": distance,
                "elevation": elevation_along_line(profile, distance),
                "geometry": transform(transform_3005_4326, line)
            }
            lake_features.append(lake_data)

    return stream_features + lake_features

from typing import List, Dict
import logging
import datetime
import openpyxl
import math
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from starlette.responses import Response
from api.v1.aggregator.schema import LayerResponse
from shapely.geometry import Point, LineString, shape
from shapely.ops import transform, nearest_points
from api.v1.aggregator.helpers import transform_3005_4326, transform_4326_3005
from api.v1.wells.helpers import distance_from_line, compass_direction_point_to_line

logger = logging.getLogger('well export')


def cross_section_xlsx_export(features: List[LayerResponse], coordinates: list, buffer: int):
    """
    packages features into an excel workbook.  Returns an HTTP response object that has the saved workbook
    ready to be returned to the client (e.g. the calling http handler can return this object directly)
    """

    workbook = openpyxl.Workbook()
    ds = workbook.active
    ds.title = "details"

    font_title = Font(size=20, bold=True, color='44546a')
    font_label = Font(bold=True)
    border_bottom = Border(bottom=Side(border_style="thick", color='4472c4'))

    # styling config
    ds['A1'].font = font_title
    ds.column_dimensions['A'].width = 20
    ds['A1'].border = border_bottom
    ds['B1'].border = border_bottom
    ds['C1'].border = border_bottom

    ds['A2'].font = font_label
    ds['A3'].font = font_label
    ds['A4'].font = font_label
    ds['A5'].font = font_label

    ds['A1'] = 'Cross section'
    ds['A2'] = 'Date generated:'
    ds['A3'] = 'A point coordinates:'
    ds['A4'] = 'B point coordinates:'
    ds['A5'] = 'Buffer radius (m):'

    cur_date = datetime.datetime.now().strftime("%X-%Y-%m-%d")

    ds['B2'] = cur_date
    ds['B3'] = str(coordinates[0][1]) + ', ' + str(coordinates[0][0])
    ds['B4'] = str(coordinates[1][1]) + ', ' + str(coordinates[1][0])
    ds['B5'] = buffer

    # create data sheets
    well_sheet = workbook.create_sheet("well")
    lith_sheet = workbook.create_sheet("lithology")
    screen_sheet = workbook.create_sheet("screen")

    # data sheet headers added
    well_sheet.append(WELL_HEADERS)
    lith_sheet.append(LITHOLOGY_HEADERS)
    screen_sheet.append(SCREEN_HEADERS)

    line = LineString([coords[:2] for coords in list(coordinates)])

    for dataset in features:
        # avoid trying to process layers if they have no features.
        if not dataset.geojson:
            continue

        # set row information
        try:
            props = dataset.geojson.features[0].properties
            well_tag_number = props["well_tag_number"]

            point = Point(shape(dataset.geojson.features[0].geometry))
            well_offset = round(distance_from_line(line, point), 2)
            well_offset_direction = compass_direction_point_to_line(line, point)

            well_values = [props.get(x, None) for x in WELL_HEADERS[3:]]
            well_sheet.append([well_tag_number] + [well_offset] \
              + [well_offset_direction] + well_values)

            lith_set = props["lithologydescription_set"]

            for item in lith_set:
                lith_values = [item.get(x, None) for x in LITHOLOGY_INDEX]
                lith_sheet.append([well_tag_number] + lith_values)

            screen_set = props["screen_set"]

            for item in screen_set:
                screen_values = [item.get(x, None) for x in SCREEN_INDEX]
                screen_sheet.append([well_tag_number] + screen_values)

        except Exception as e:
            logger.warn(e)
            continue

    # set header style for data sheets
    set_row_style(well_sheet)
    set_row_style(lith_sheet)
    set_row_style(screen_sheet)

    # fix column widths
    set_column_width(well_sheet)
    set_column_width(lith_sheet)
    set_column_width(screen_sheet)

    filename = f"{cur_date}_CrossSection"

    response = Response(
        content=save_virtual_workbook(workbook),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}.xlsx'})

    return response


def set_column_width(sheet):
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length


def set_row_style(sheet):
    font_header = Font(bold=True, color='FFFFFF')
    fill_header = PatternFill("solid", fgColor="4472c4")
    for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            cell.font = font_header
            cell.fill = fill_header


SCREEN_INDEX = [
    'start',
    'end',
    'diameter',
    'assembly_type',
    'slot_size'
]

SCREEN_HEADERS = [
    'well_tag_number',
    'screen_from',
    'screen_to',
    'screen_diameter',
    'screen_assembly_type',
    'screen_slot_size'
]

LITHOLOGY_INDEX = [
    'start',
    'end',
    'lithology_raw_data',
    'lithology_description',
    'lithology_material',
    'lithology_hardness',
    'lithology_colour',
    'water_bearing_estimated_flow',
    'lithology_observation'
]

LITHOLOGY_HEADERS = [
    'well_tag_number',
    'lithology_from',
    'lithology_to',
    'lithology_raw_data',
    'lithology_description_code',
    'lithology_material_code',
    'lithology_hardness_code',
    'lithology_colour_code',
    'water_bearing_estimated_flow',
    'lithology_observation'
]

WELL_HEADERS = [
    'well_tag_number',
    "well_offset",
    "well_offset_direction",
    "identification_plate_number",
    "well_identification_plate_attached",
    "well_status",
    "well_class",
    "well_subclass",
    "intended_water_use",
    "licenced_status",
    "observation_well_number",
    "obs_well_status",
    "water_supply_system_name",
    "water_supply_system_well_name",
    "street_address",
    "city",
    "legal_lot",
    "legal_plan",
    "legal_district_lot",
    "legal_block",
    "legal_section",
    "legal_township",
    "legal_range",
    "land_district",
    "legal_pid",
    "well_location_description",
    "latitude",
    "longitude",
    "utm_zone_code",
    "utm_northing",
    "utm_easting",
    "coordinate_acquisition_code",
    "construction_start_date",
    "construction_end_date",
    "alteration_start_date",
    "alteration_end_date",
    "decommission_start_date",
    "decommission_end_date",
    "diameter",
    "total_depth_drilled",
    "finished_well_depth",
    "bedrock_depth",
    "final_casing_stick_up",
    "ground_elevation",
    "ground_elevation_method",
    "static_water_level",
    "well_yield",
    "well_yield_unit",
    "artesian_flow",
    "artesian_pressure",
    "comments",
    "ems",
    "aquifer",
    "aquifer_vulnerability_index",
    "storativity",
    "transmissivity",
    "hydraulic_conductivity",
    "specific_storage",
    "specific_yield",
    "testing_method",
    "testing_duration",
    "analytic_solution_type",
    "boundary_effect",
    "aquifer_lithology",
    "well_publication_status",
]

WELL_NUMBER_COLUMNS = [
    "diameter",
    "finished_well_depth",
    "bedrock_depth",
    "ground_elevation",
    "static_water_level",
    "well_yield",
    "artesian_flow"
]

WELLS_NEARBY_HEADERS = [
    "well_tag_number",
    "latitude",
    "longitude",
    "well_yield",
    "diameter",
    "well_yield_unit",
    "finished_well_depth",
    "street_address",
    "intended_water_use",
    "aquifer_subtype",
    "aquifer_hydraulically_connected",
    "static_water_level",
    "top_of_screen",
    "top_of_screen_type",
    "distance",
    "swl_to_screen",
    "swl_to_bottom_of_well",
    "aquifer_id",
    "aquifer_material",
    "aquifer_lithology",
]


def wells_by_aquifer_xlsx_export(wells_by_aquifer: Dict):
    """
    Creates an excel file with a list of wells separated by aquifers
    Each aquifer has its on worksheet with all the corresponding wells
    Returns a response object with the excel data as content
    """

    aquifer_count = len(wells_by_aquifer)

    # There is no data
    if aquifer_count <= 0:
        return None

    workbook = openpyxl.Workbook()

    # A workbook is automatically created with 1 sheet which is set as the active one.
    # First aquifer wells goes into this sheet.
    aquifer_1, aquifer_1_wells = next(iter(wells_by_aquifer.items()))
    sheet1 = workbook.active
    sheet1.title = f"Aquifer {aquifer_1}"

    # Helper function to get row information
    def get_well_values(a_well, headers):
        well_dict = dict(a_well)

        if a_well.aquifer:
            del well_dict['aquifer']
            well_dict['aquifer_id'] = a_well.aquifer.aquifer_id
            well_dict['aquifer_material'] = a_well.aquifer.material_desc

        well_values = [well_dict.get(x, None) for x in headers]
        return well_values

    sheet1.append(WELLS_NEARBY_HEADERS)
    for well in wells_by_aquifer[aquifer_1]:
        sheet1.append(get_well_values(well, WELLS_NEARBY_HEADERS))

    sheets = {}

    # Create worksheets
    if aquifer_count > 1:
        for aquifer in list(wells_by_aquifer)[1:]:

            # Categorize unknown aquifers as 'Other'
            other = 'Uncorrelated'
            aquifer_sheet = f"Aquifer {aquifer}" if aquifer else other

            sheets[aquifer_sheet] = workbook.create_sheet(aquifer_sheet)
            sheets[aquifer_sheet].append(WELLS_NEARBY_HEADERS)
            for well in wells_by_aquifer[aquifer]:
                sheets[aquifer_sheet].append(get_well_values(well, WELLS_NEARBY_HEADERS))

    cur_date = datetime.datetime.now().strftime("%X-%Y-%m-%d")

    filename = f"{cur_date}_WellsNearby"
    response = Response(
        content=save_virtual_workbook(workbook),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}.xlsx'})

    return response


import math
import pyproj
from shapely.ops import transform, nearest_points
from shapely.geometry import Point, LineString
from api.v1.aggregator.helpers import transform_4326_3005, transform_3005_4326

COMPASS_BRACKETS = ["N", "NE", "E", "SE", "S", "SW", "W", "NW", "N"]


def distance_from_line(line: LineString, point: Point, srid=4326):
    """
    calculates the shortest distance between the point and line.
    """
    if srid == 4326:
        # transform to BC Albers, which has a base unit of metres
        point = transform(transform_4326_3005, point)
        line = transform(transform_4326_3005, line)

    elif srid != 3005:
        raise ValueError("SRID must be either 4326 or 3005")

    return point.distance(line)


def compass_direction_point_to_line(line: LineString, point: Point, srid=4326):
    """
    calculates the azimuth from a point along a line to a point.
    angle then used to lookup the compass direction
    """
    if srid == 4326:
        # transform to BC Albers, which has a base unit of metres
        point = transform(transform_4326_3005, point)
        line = transform(transform_4326_3005, line)

    elif srid != 3005:
        raise ValueError("SRID must be either 4326 or 3005")

    nearest_point = nearest_points(line, point)[0]
    angle = math.atan2(point.x - nearest_point.x, point.y - nearest_point.y)
    degrees = math.degrees(angle) if angle >= 0 else math.degrees(angle) + 360
    compass_lookup = round(degrees / 45)
    compass_direction = COMPASS_BRACKETS[compass_lookup]

    return compass_direction


import json
from typing import List
from logging import getLogger
from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy.orm import Session
from shapely.geometry import Point, LineString, mapping

from api.db.utils import get_db
from api.v1.wells.excel import wells_by_aquifer_xlsx_export
from api.v1.wells.schema import WellDrawdown, CrossSection, CrossSectionExport, WellsExport
from api.v1.elevations.controllers.profile import get_profile_line_by_length
from api.v1.elevations.controllers.surface import fetch_surface_lines
from api.v1.wells.controller import (
    get_waterbodies_along_line,
    create_line_buffer,
    get_wells_with_drawdown,
    get_wells_by_aquifer,
    get_wells_along_line,
    get_parallel_line_offset,
    get_cross_section_export
)

logger = getLogger("wells")

router = APIRouter()


@router.get("/nearby", response_model=List[WellDrawdown])
def get_nearby_wells(
        point: str = Query(..., title="Point of interest",
                           description="Point of interest to centre search at"),
        radius: float = Query(1000, title="Search radius",
                              description="Search radius from point", ge=0, le=10000),
):
    """ finds wells near a given point
        gets wells and their corresponding subsurface data from GWELLS
        and computes distance
    """
    point_parsed = json.loads(point)
    point_shape = Point(point_parsed)

    wells_nearby = get_wells_with_drawdown(point_shape, radius)

    return wells_nearby


@router.get("/nearby/aquifers")
def get_nearby_wells_by_aquifer(
        point: str = Query(..., title="Point of interest",
                           description="Point of interest to centre search at"),
        radius: float = Query(1000, title="Search radius",
                              description="Search radius from point", ge=0, le=10000),
):
    """ finds wells near to a point
        fetches distance data using the Wally database, and combines
        it with subsurface data from GWELLS
    """
    point_parsed = json.loads(point)
    point_shape = Point(point_parsed)

    return get_wells_by_aquifer(point_shape, radius)


@router.post("/nearby/export")
def export_nearby_wells(
        req: WellsExport
):
    """ 
    finds wells near to a point
    fetches distance data using the Wally database,
    combines it with screen data from GWELLS and
    filters based on well list in request
    """
    point_parsed = json.loads(req.point)
    export_wells = req.export_wells

    well_tag_numbers = ','.join([str(wtn) for wtn in export_wells])
    point_shape = Point(point_parsed)

    wells_by_aquifer = get_wells_by_aquifer(point_shape, req.radius, well_tag_numbers)

    return wells_by_aquifer_xlsx_export(wells_by_aquifer)


@router.get("/section", response_model=CrossSection)
def get_wells_section(
        db: Session = Depends(get_db),
        line: str = Query(..., title="Section line",
                          description="Section line along which wells will be plotted"),
        radius: float = Query(200, title="Search radius",
                              description="Search radius (or offset) from line", ge=0, le=10000)
):
    """ search for wells along a line, returning a cross section of well data """

    line_parsed = json.loads(line)
    line_shape = LineString(line_parsed)

    left = get_parallel_line_offset(db, line_shape, -radius)
    left_half = get_parallel_line_offset(db, line_shape, -radius / 2)
    right = get_parallel_line_offset(db, line_shape, radius)
    right_half = get_parallel_line_offset(db, line_shape, radius / 2)
    lines = [left[0], left_half[0], line_shape.wkt, right_half[0], right[0]]

    # surface of 5 lines used for 3d display
    try:
        surface = fetch_surface_lines(lines)
    except:
        raise HTTPException(
            status_code=502, detail="unable to retrieve elevations from GeoGratis CDEM API")

    profile_line_linestring = surface[2]
    profile_line = get_profile_line_by_length(db, profile_line_linestring)
    wells_along_line = get_wells_along_line(
        db, profile_line_linestring, radius)

    buffer = create_line_buffer(profile_line_linestring, radius)

    surface_lines = [list(line.coords) for line in surface]
    # we need to reverse the point lists for the -radius results
    surface_lines[0].reverse()
    surface_lines[1].reverse()

    # waterbodies that cross profile
    waterbodies_along_line = get_waterbodies_along_line(
        line_shape, profile_line_linestring)

    # logger.info(surface_lines)
    section = CrossSection(search_area=mapping(buffer), wells=wells_along_line,
                           waterbodies=waterbodies_along_line,
                           elevation_profile=profile_line, surface=surface_lines)

    return section


@router.post("/section/export")
def get_wells_section_export(
        req: CrossSectionExport
):
    """ gather well information for many wells and export an excel report """

    return get_cross_section_export(req)



import json

import geojson
import logging

from geojson import Point, Feature, FeatureCollection
from shapely import wkb, wkt
from shapely.geometry import LineString, CAP_STYLE, JOIN_STYLE, shape, mapping, Point
from shapely.ops import transform, split, snap
from typing import List
from sqlalchemy import text, func
from sqlalchemy.orm import Session
from api.layers.freshwater_atlas_stream_networks import FreshwaterAtlasStreamNetworks
from api.v1.streams.schema import StreamPoint
from api.v1.aggregator.controller import feature_search
from api.v1.aggregator.helpers import transform_3005_4326, transform_4326_3005
from api.layers.freshwater_atlas_stream_networks import FreshwaterAtlasStreamNetworks

logger = logging.getLogger("api")


def get_nearest_streams_by_ogc_fid(db: Session, search_point: Point, ogc_fids: list) -> list:
    streams_q = db.query(
        FreshwaterAtlasStreamNetworks.OGC_FID.label("ogc_fid"),
        FreshwaterAtlasStreamNetworks.LENGTH_METRE.label("length_metre"),
        FreshwaterAtlasStreamNetworks.FEATURE_SOURCE.label("feature_source"),
        FreshwaterAtlasStreamNetworks.GNIS_NAME.label("gnis_name"),
        FreshwaterAtlasStreamNetworks.LEFT_RIGHT_TRIBUTARY.label(
            "left_right_tributary"),
        FreshwaterAtlasStreamNetworks.GEOMETRY_LEN.label("geometry_length"),
        FreshwaterAtlasStreamNetworks.WATERSHED_GROUP_CODE.label(
            "watershed_group_code"),
        FreshwaterAtlasStreamNetworks.FWA_WATERSHED_CODE.labe(
            "fwa_watershed_code"),
        func.ST_ASText(FreshwaterAtlasStreamNetworks.GEOMETRY).label(
            "geometry"),
        # FreshwaterAtlasStreamNetworks,
        func.ST_Distance(
            FreshwaterAtlasStreamNetworks.GEOMETRY,
            func.ST_SetSRID(func.ST_GeomFromText(search_point.wkt), 4326)
        ).label('distance_degrees'),
        func.ST_Distance(
            func.Geography(FreshwaterAtlasStreamNetworks.GEOMETRY),
            func.ST_GeographyFromText(search_point.wkt)
        ).label('distance'),
        func.ST_AsGeoJSON(func.ST_ClosestPoint(
            FreshwaterAtlasStreamNetworks.GEOMETRY,
            func.ST_SetSRID(func.ST_GeomFromText(search_point.wkt), 4326))
        ).label('closest_stream_point')
    ).filter(FreshwaterAtlasStreamNetworks.OGC_FID.in_(ogc_fids))
    logging.debug(streams_q)

    rs_streams = streams_q.all()
    columns = [col['name'] for col in streams_q.column_descriptions]

    streams_with_columns = []
    for row in rs_streams:
        stream = {columns[i]: item for i, item in enumerate(row)}
        stream['geojson'] = get_feature_geojson(stream)
        stream['closest_stream_point'] = json.loads(
            stream['closest_stream_point'])
        streams_with_columns.append(stream)

    logging.debug(streams_with_columns)
    return streams_with_columns


def get_streams_with_apportionment(
        db: Session,
        search_point: Point,
        limit=10,
        get_all=False,
        with_apportionment=True,
        weighting_factor=2) -> list:
    streams = get_nearest_streams(db, search_point, limit)

    if not with_apportionment:
        return streams

    streams_with_apportionment = get_apportionment(
        streams, weighting_factor, get_all)
    return streams_with_apportionment


def get_nearest_streams(db: Session, search_point: Point, limit=10) -> list:
    # Get the nearest 10 streams to the point
    sql = text("""
      WITH nearest_streams AS (
          select    *
          from      freshwater_atlas_stream_networks streams
          order by  streams."GEOMETRY" <#>
                        ST_SetSRID(ST_GeomFromText(:search_point), 4326)
          limit     10
      )
      SELECT 
        nearest_streams."OGC_FID" as id, 
        nearest_streams."OGC_FID" as ogc_fid,
        nearest_streams."LENGTH_METRE" as length_metre,
        nearest_streams."FEATURE_SOURCE" as feature_source,
        nearest_streams."GNIS_NAME" as gnis_name,
        nearest_streams."LINEAR_FEATURE_ID" as linear_feature_id,
        nearest_streams."LEFT_RIGHT_TRIBUTARY" as left_right_tributary,
        nearest_streams."GEOMETRY.LEN" as geometry_length,
        ST_AsText(nearest_streams."GEOMETRY") as geometry,
        nearest_streams."WATERSHED_GROUP_CODE" as watershed_group_code, 
        nearest_streams."FWA_WATERSHED_CODE" as fwa_watershed_code,
        ST_Distance(nearest_streams."GEOMETRY",
          ST_SetSRID(ST_GeomFromText(:search_point), 4326)
        ) AS distance_degrees,
        ST_Distance(nearest_streams."GEOMETRY"::geography,
          ST_SetSRID(ST_GeographyFromText(:search_point), 4326)
        ) AS distance,
        ST_AsText(ST_SetSRID(ST_GeomFromText(:search_point), 4326)) as search_point,
        ST_AsGeoJSON(ST_ClosestPoint(
        nearest_streams."GEOMETRY", 
        ST_SetSRID(ST_GeomFromText(:search_point), 4326))) as closest_stream_point
      FROM      nearest_streams
      ORDER BY  ST_Distance(nearest_streams."GEOMETRY", ST_SetSRID(ST_GeomFromText(:search_point), 4326)) ASC
      LIMIT     :limit 
    """)
    rp_nearest_streams = db.execute(
        sql, {'search_point': search_point.wkt, 'limit': limit})
    nearest_streams = [
        dict(row,
             geojson=get_feature_geojson(row),
             closest_stream_point=json.loads(row['closest_stream_point'])
             ) for row in rp_nearest_streams
    ]

    return nearest_streams


def get_apportionment(streams, weighting_factor, get_all=False, force_recursion=False):
    """Recursive function that gets the apportionment (in percentage) for all streams"""

    logger.debug('get_apportionment')
    # Don't do recursion if there are more than 10 streams
    if len(streams) > 10 and not get_all and not force_recursion:
        raise RecursionError('Cannot compute apportionment for more than 10 streams. Set '
                             'force_recursion=True.')

    # Get the summation of the inverse distance formula
    total = 0
    for stream in streams:
        stream['inverse_distance'] = get_inverse_distance(
            stream['distance'], weighting_factor)
        total += stream['inverse_distance']

    # We need to loop again after we have the total so we know the percentage
    for i, stream in enumerate(streams):
        percentage = (stream['inverse_distance'] / total) * 100
        # Delete stream if apportionment is less than 10% and re-calculate
        # Skip this if get_all is True
        if percentage < 10 and not get_all:
            del streams[i]
            return get_apportionment(streams, weighting_factor)
        stream['apportionment'] = percentage

    return streams


def get_inverse_distance(stream_distance, weighting_factor):
    return 1 / (stream_distance ** weighting_factor)


def get_feature_geojson(stream) -> Feature:
    stream_copy = dict(stream)
    del stream_copy['closest_stream_point']
    del stream_copy['ogc_fid']
    feature = Feature(
        geometry=wkt.loads(stream_copy['geometry']),
        id=stream['ogc_fid'],
        properties=dict(stream_copy)
    )
    del stream_copy['geometry']
    return feature


def get_connected_streams(db: Session, outflowCode: str) -> list:
    q = db.query(FreshwaterAtlasStreamNetworks) \
        .filter(FreshwaterAtlasStreamNetworks.FWA_WATERSHED_CODE.startswith(outflowCode))

    results = q.all()

    feature_results = [FreshwaterAtlasStreamNetworks.get_as_feature(
        row, FreshwaterAtlasStreamNetworks.GEOMETRY) for row in results]

    return feature_results


def get_nearest_hydat_stream_segments(db: Session, station_number: str) -> List[StreamPoint]:
    """ get nearest stream segments returns the 5 stream segments nearest the
        HYDAT station that match the name of the station.

        This helps with inferring the station location (in cases where the location
        is not clear) because we can run a watershed delineation from each segment
        and compare the results with the listed drainage_area_gross to infer which
        segment the station is likely on.
    """

    q = """
      WITH stn AS (
        select station_name, geom
        from hydat.stations
        where station_number = :station_number
      ),
      nearest_streams AS (
          select    *
          from      freshwater_atlas_stream_networks streams
          order by  streams."GEOMETRY" <#>
                        (select geom from stn)
          limit     5
      )
      SELECT 
        (select ST_AsBinary(geom) from stn) as station_point,
        nearest_streams."GNIS_NAME" as gnis_name,
        nearest_streams."LINEAR_FEATURE_ID" as linear_feature_id,
        ST_AsBinary(
          ST_ClosestPoint(
            nearest_streams."GEOMETRY", 
            (select geom from stn)
          )
        ) as stream_point
      FROM      nearest_streams
      ORDER BY  
        coalesce(nearest_streams."GNIS_NAME" ILIKE '%' || (select split_part(station_name, ' ', 1) from stn) || '%', FALSE) DESC,
        ST_Distance(nearest_streams."GEOMETRY", (select geom from stn)) ASC
    """
    res = db.execute(q, {"station_number": station_number})
    streams = []
    for row in res:
        row = dict(row)
        streams.append(StreamPoint(
            stream_point=wkb.loads(row['stream_point'].tobytes()),
            stream_feature_id=row['linear_feature_id']
        ))
    return streams


def get_connected_streams(db: Session, outflowCode: str) -> list:
    q = db.query(FreshwaterAtlasStreamNetworks).filter(
        FreshwaterAtlasStreamNetworks.FWA_WATERSHED_CODE.startswith(outflowCode))

    results = q.all()

    feature_results = [FreshwaterAtlasStreamNetworks.get_as_feature(
        row, FreshwaterAtlasStreamNetworks.GEOMETRY) for row in results]

    return feature_results


def watershed_root_code(code: str) -> str:
    """ truncates zero values from the end of watershed codes, .
        Watershed codes have 21 segments, but most of the segments will
        be a "zero value". E.g.:
        100-123333-432123-000000-000000-000000- .... etc
    """

    return [x for x in code.split('-') if x != '000000']


def to_3005(from_proj, feat):
    """ returns a feature in EPSG:3005, transforming if necessary
        `from_proj` supports 4326 and 3005.
    """

    if from_proj == 3005:
        return feat

    elif from_proj == 4326:
        return transform(transform_4326_3005, feat)

    logger.warn(
        'to_3005: from_proj must be either 4326 or 3005. Feature returned without transforming to 3005.')
    return feat


def get_closest_stream_segment(db: Session, point: Point):
    sql = text("""
      SELECT
        streams."GNIS_NAME" as gnis_name,
        streams."LINEAR_FEATURE_ID" as linear_feature_id,
        streams."DOWNSTREAM_ROUTE_MEASURE" as downstream_route_measure,
        ST_AsGeoJSON(ST_ClosestPoint(
        streams."GEOMETRY", 
        ST_SetSRID(ST_GeomFromText(:search_point), 4326))) as closest_stream_point
      FROM
      freshwater_atlas_stream_networks as streams
      ORDER BY streams."GEOMETRY" <->
        ST_SetSRID(ST_GeomFromText(:search_point), 4326)
      LIMIT 1
    """)
    segment = db.execute(sql, {'search_point': point.wkt}).fetchone()
    return dict(segment,
                closest_stream_point=json.loads(segment.closest_stream_point))


def split_line_by_closest_point(
        line: LineString,
        point: Point):

    distance = line.project(transform(transform_4326_3005, point))
    interpolated_point = line.interpolate(distance)
    snap_line = snap(line, interpolated_point, 0.001)
    split_lines = split(snap_line, interpolated_point)

    return split_lines


def get_stream_line(
        db: Session,
        linear_feature_id: int):

    return db.execute(
        """
        select ST_AsGeoJSON(ST_Transform("GEOMETRY", 3005))
        as "GEOMETRY" from freshwater_atlas_stream_networks
        where "LINEAR_FEATURE_ID" = :linear_feature_id
        """,
        {
            "linear_feature_id": linear_feature_id
        }).fetchone()


def get_split_line_stream_buffers(
        db: Session,
        linear_feature_id: int,
        buffer: float,
        point: Point):

    db_line = get_stream_line(db, linear_feature_id)
    segment = shape(geojson.loads(db_line[0]) if db_line[0] else None)
    split_lines = split_line_by_closest_point(segment, point)
    buffer_lines = [transform(transform_3005_4326, line.buffer(buffer)) for line in split_lines]

    return buffer_lines


def get_downstream_area(
        db: Session,
        linear_feature_id: int,
        buffer: float):

    q = """
        with watershed_code_stats as (
            SELECT DISTINCT
                "FWA_WATERSHED_CODE" as fwa_code,
                "LOCAL_WATERSHED_CODE" as loc_code,
                "DOWNSTREAM_ROUTE_MEASURE" as downstream_route_measure,
                left(
                    regexp_replace(
                        "FWA_WATERSHED_CODE",
                        '000000',
                        '%'
                    ),
                    strpos(regexp_replace("FWA_WATERSHED_CODE", '000000', '%'),
                    '%'
                )) as fwa_prefix
            FROM freshwater_atlas_stream_networks
            WHERE   "LINEAR_FEATURE_ID" = :linear_feature_id
        ),
        selected_stream as (
            select  ST_Transform(
                ST_Buffer(
                    ST_Transform("GEOMETRY", 3005),
                    :buffer),
                    4326
                ) as "GEOMETRY" from freshwater_atlas_stream_networks, watershed_code_stats
            where "FWA_WATERSHED_CODE" = fwa_code
            and
            (CASE
              WHEN "DOWNSTREAM_ROUTE_MEASURE" = 0
              THEN "DOWNSTREAM_ROUTE_MEASURE" <= watershed_code_stats.downstream_route_measure
              ELSE "DOWNSTREAM_ROUTE_MEASURE" < watershed_code_stats.downstream_route_measure
            END)
        )
        select
            ST_AsGeoJSON(ST_Union("GEOMETRY"))
        from    (
            select ST_MakeValid("GEOMETRY") "GEOMETRY" from watershed_code_stats, selected_stream
        ) subq   
        """

    return db.execute(
        q,
        {
            "linear_feature_id": linear_feature_id,
            "buffer": buffer,
        }).fetchone()


def get_upstream_area(
        db: Session,
        linear_feature_id: int,
        buffer: float,
        full_upstream_area: bool):
    """ returns the polygon area upstream from the selected stream feature
    (using the linear_feature_id property of a Freshwater Atlas Stream Networks stream segment) """

    # Gather up the selected stream segments (from the stream's own headwaters
    # down to the mouth of the stream where it drains into the next river),
    # as well as all *upstream* tributary networks from the selected reach.
    # This represents the entire drainage network upstream of the selected
    # reach, combined with just the stream's own geometry downstream (no
    # tributaries downstream of the selected reach are included).
    #
    # these queries work by inspecting the last non-zero code of the local
    # watershed code, which roughly represents the percent distance along the
    # stream of each segment of the stream.

    # First default to the full upstream area by using Freshwater Atlas Watershed polygons that match the
    # stream network codes. This is faster than creating a buffer from all the stream segments,
    # but only works well upstream from the point of interest. This produces a large, filled in
    # polygon that follows the shape of the drainage basin.
    q = """
    with watershed_code_stats as (
        SELECT DISTINCT
            "FWA_WATERSHED_CODE" as fwa_code,
            "LOCAL_WATERSHED_CODE" as loc_code,
            "DOWNSTREAM_ROUTE_MEASURE" as downstream_route_measure,
            (FLOOR(((strpos(regexp_replace("LOCAL_WATERSHED_CODE", '000000', '%'), '%')) - 4) / 7) + 1)::int
                as loc_code_last_nonzero_code,
            left(
                regexp_replace(
                    "FWA_WATERSHED_CODE",
                    '000000',
                    '%'
                ),
                strpos(regexp_replace("FWA_WATERSHED_CODE", '000000', '%'),
                '%'
            )) as fwa_prefix
        FROM freshwater_atlas_stream_networks
        WHERE   "LINEAR_FEATURE_ID" = :linear_feature_id
    ),
    streams as (
        select ST_Transform(
            ST_Buffer(
                ST_Transform("GEOMETRY", 3005),
                :buffer),
                4326
            ) as "GEOMETRY" from freshwater_atlas_stream_networks, watershed_code_stats
        where   "FWA_WATERSHED_CODE" = fwa_code and "DOWNSTREAM_ROUTE_MEASURE" >= watershed_code_stats.downstream_route_measure
    )
    select
        ST_AsGeoJSON(ST_Union("GEOMETRY"))
    from    (
        select ST_MakeValid("GEOMETRY") "GEOMETRY" from streams 
        union all
        select  ST_MakeValid("GEOMETRY") "GEOMETRY" from freshwater_atlas_watersheds, watershed_code_stats
        where   "FWA_WATERSHED_CODE" like fwa_prefix
        AND     split_part(
                    "LOCAL_WATERSHED_CODE", '-',
                    watershed_code_stats.loc_code_last_nonzero_code
                )::int > split_part(
                    watershed_code_stats.loc_code, '-',
                    watershed_code_stats.loc_code_last_nonzero_code
                )::int
    ) subq   
    """

    # if the user overrides searching within the full upstream catchment area, search only within <buffer>
    # metres of the stream. This produces a polygon with narrow branches that follows the shape of
    # the stream network.
    # The main difference between this query and the default query is that the "from" subquery selects
    # from FWA Stream Networks, returning buffered linestrings, which takes longer.
    if not full_upstream_area:
        q = """
        with watershed_code_stats as (
            SELECT
                "FWA_WATERSHED_CODE" as fwa_code,
                "LOCAL_WATERSHED_CODE" as loc_code,
                "DOWNSTREAM_ROUTE_MEASURE" as downstream_route_measure,
                (FLOOR(((strpos(regexp_replace("LOCAL_WATERSHED_CODE", '000000', '%'), '%')) - 4) / 7) + 1)::int
                    as loc_code_last_nonzero_code,
                left(
                    regexp_replace("FWA_WATERSHED_CODE", '000000', '%'),
                    strpos(regexp_replace("FWA_WATERSHED_CODE", '000000', '%'),
                    '%')
                ) as fwa_prefix
            FROM freshwater_atlas_stream_networks
            WHERE   "LINEAR_FEATURE_ID" = :linear_feature_id
        )
        select
            ST_AsGeoJSON(
                ST_Transform(
                    ST_Buffer(
                        ST_Transform(ST_Collect("GEOMETRY"), 3005),
                        :buffer, 'endcap=round join=round'
                    ),
                    4326
                )
            )
        from    (
            select  "GEOMETRY" from freshwater_atlas_stream_networks, watershed_code_stats
            where   "FWA_WATERSHED_CODE" = fwa_code and "DOWNSTREAM_ROUTE_MEASURE" >= downstream_route_measure
            union all
            select  "GEOMETRY" from freshwater_atlas_stream_networks, watershed_code_stats
            where   "FWA_WATERSHED_CODE" like fwa_prefix
            AND     
            (CASE
              WHEN watershed_code_stats.downstream_route_measure = 0
              THEN 
                split_part(
                  "LOCAL_WATERSHED_CODE", '-',
                  watershed_code_stats.loc_code_last_nonzero_code + 1
                )::int > split_part(
                    watershed_code_stats.loc_code, '-',
                    watershed_code_stats.loc_code_last_nonzero_code + 1
                )::int
              ELSE 
                split_part(
                  "LOCAL_WATERSHED_CODE", '-',
                  watershed_code_stats.loc_code_last_nonzero_code
                )::int > split_part(
                    watershed_code_stats.loc_code, '-',
                    watershed_code_stats.loc_code_last_nonzero_code
                )::int
            END)
        ) subq
        """

    return db.execute(
        q,
        {
            "linear_feature_id": linear_feature_id,
            "buffer": buffer,
        }).fetchone()


def get_features_within_buffer(db: Session, line, distance: float, layer: str) -> FeatureCollection:
    """ List features within a buffer zone from a geometry
    """
    if not line:
        return None

    buf_simplified = line.minimum_rotated_rectangle

    fc = feature_search(db, [layer], buf_simplified)[0].geojson

    features = fc.get('features')
    feat_proj = 4326

    # indicate if features are BC Albers projection
    # DataBC feature collections have a `crs` property containing the
    # name `urn:ogc:def:crs:EPSG::3005`
    if fc.get('crs') and fc.get('crs', {}).get('properties', {}).get('name', "").endswith("3005"):
        feat_proj = 3005

    line_3005 = transform(transform_4326_3005, line)

    features_intersecting = [
        Feature(
            geometry=feat['geometry'],
            properties=feat['properties']) for feat in features if to_3005(feat_proj, shape(feat['geometry'])
                                                                           ).intersects(line_3005)
    ]

    return FeatureCollection(features_intersecting)

def databc_feature_search(layer, search_area=None, cql_filter=None) -> FeatureCollection:
    """ looks up features from `layer` in `search_area`.
        Layer should be in DATABC_LAYER_IDS.
        Search area should be SRID 4326.
    """
    if not search_area and not cql_filter:
        raise HTTPException(
            status_code=400, detail="Must provide either search_area or cql_filter")

    if search_area and cql_filter:
        raise HTTPException(
            status_code=400, detail="Must provide either search_area or cql_filter, not both")

    if search_area:
        search_area = transform(transform_4326_3005, search_area)
        cql_filter = f"""
                INTERSECTS({DATABC_GEOMETRY_FIELD.get(
                    layer, 'GEOMETRY')}, {search_area.wkt})
            """

    query = WMSGetFeatureQuery(
        typeName=DATABC_LAYER_IDS.get(
            layer, layer),
        cql_filter=cql_filter
    )

    req = ExternalAPIRequest(
        url=f"https://openmaps.gov.bc.ca/geo/pub/wfs?",
        layer=layer,
        q=query
    )
    feature_list = fetch_geojson_features([req])

    if not len(feature_list):
        raise HTTPException(status_code=404, detail="Dataset not found")

    return feature_list[0].geojson
